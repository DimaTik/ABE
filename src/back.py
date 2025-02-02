import bs4
import requests
import openpyxl as xl
import datetime
import pprint
import time
from openpyxl.styles import Alignment
import openpyxl.utils
import getpass

months = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь', 'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']


class Consultant:
	def __init__(self, year):
		self.quarters = {'Январь': [0, 0], 'Февраль': [0, 1], 'Март': [0, 2], 'Апрель': [1, 0], 'Май': [1, 1],
						 'Июнь': [1, 2], 'Июль': [2, 0], 'Август': [2, 1], 'Сентябрь': [2, 2], 'Ноябрь': [3, 0],
						 'Октябрь': [3, 1], 'Декабрь': [3, 2]}
		self.duration = {
			'40': 0,
			'36': 1,
			'24': 2
		}
		headers = {
			'Accept': "*/*",
			'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36 OPR/115.0.0.0 (Edition Yx 05)"
		}
		user = getpass.getuser()
		url = f'https://www.consultant.ru/law/ref/calendar/proizvodstvennye/{year}'
		req = requests.get(url, headers)
		with open(fr'C:\Users\{user}\AppData\Local\Temp\calendar_{year}', 'w', encoding='utf-8') as file:
			file.write(req.text)
		with open(fr'C:\Users\{user}\AppData\Local\Temp\calendar_{year}', 'r', encoding='utf-8') as file:
			src = file.read()
		self.soup = bs4.BeautifulSoup(src, 'lxml')

	def get_standard_hours(self, month):
		quarter = self.quarters[month][0]
		quarter_list = self.soup.find('div', class_='block-print').find_next_sibling().find_next_sibling(). \
			find_all('div', class_='row')  # Нашли divs
		quarter_list_result = [quarter_list[4]]  # Нашли первый квартал
		quarter_list = quarter_list[5:]  # Обрезали ввиду верстки
		for i in range(5, len(quarter_list), 6):
			quarter_list_result.append(quarter_list[i])
		standard_hours = \
			float(
				quarter_list_result[quarter].find_all_next('div', class_='col-md-3 col-xs-2')
				[self.quarters[month][1]].text \
					.split()[self.duration['40']].strip().replace(',', '.'))  # Магия и ты гений
		return standard_hours


class Excel:
	def __init__(self, month, path):
		self.month = month
		self.initial_workbook = xl.load_workbook(path)
		self.result_workbook = xl.Workbook()
		self.font_size = 11
		self.data = {}

	def get_jobs_and_div(self):
		sheet = self.initial_workbook.active
		for row in sheet.iter_rows(2, sheet.max_row):
			self.data[row[1].value] = [row[2].value, row[3].value, row[5].value]
		return self.data

	def create_result_table(self, data):
		sheet = self.result_workbook.create_sheet(data[0])
		for i, j in zip(range(1, 7), ('Проект', 'ФИО', 'Должность', 'Отдел', 'Часы', 'Деньги')):
			sheet.cell(row=1, column=i).value = j 	# Форматирование верхней строчки
		sheet.merge_cells(start_row=2, end_row=len(data[1])+1, start_column=1, end_column=1)
		sheet['A2'] = data[0] 	# Проект
		sheet['A2'].alignment = Alignment(horizontal='center', vertical='center')
		sheet.merge_cells(start_row=len(data[1])+2, end_row=len(data[1])+2, start_column=1, end_column=4)
		sheet[f'A{len(data[1])+2}'] = 'Итого'
		sheet[f'A{len(data[1])+2}'].alignment = Alignment(horizontal='center', vertical='center')
		self.result_workbook.save(f'{self.month}_{datetime.datetime.today().year}.xlsx')

	def set_result_table(self, data):
		sheet = self.result_workbook[data[0]]
		for i in range(len(data[1])):
			for j in range(5):
				sheet.cell(row=i+2, column=j+2).value = data[1][i][j]  # Заполнение данных
		for i in range(2, 4):
			length = max(len(str(sheet.cell(row=j, column=i).value)) for j in range(1, len(data[1])+1)) * (self.font_size**(self.font_size*0.009))
			for j in range(1, len(data[1]) + 1):
				sheet.column_dimensions[xl.utils.get_column_letter(i)].width = length
		sheet.cell(row=len(data[1])+2, column=5).value = sum([data[1][i][3] for i in range(len(data[1]))]) 	# Итого часов на проект
		sheet.cell(row=len(data[1])+2, column=6).value = sum([data[1][i][4] for i in range(len(data[1]))])  # Итого выплаченных денег
		self.result_workbook.save(f'{self.month}_{datetime.datetime.today().year}.xlsx')

	def del_sheet1(self):
		del self.result_workbook[self.result_workbook.sheetnames[0]]
		self.result_workbook.save(f'{self.month}_{datetime.datetime.today().year}.xlsx')


class Bitrix:
	def __init__(self, path, month, year):
		self.path = fr'{path}\{month}_{year}_bitrix.xlsx'
		self.workbook = xl.load_workbook(self.path)
		self.sheet = self.workbook.active

	def get_hours_worked(self):
		data = {}
		for proj in range(3, self.sheet.max_column):
			for row in range(1, self.sheet.max_row):
				data[str(int(self.sheet[2][proj].value[:4]))] = dict([(self.sheet[i][1].value, self.sheet[i][proj].value) for i in range(3, self.sheet.max_row-1)])
		return data

	def get_projects(self):
		return [int(self.sheet[2][i].value[:4]) for i in range(3, self.sheet.max_row - 3)]

	def get_month_from_name(self):
		return self.path.split('\\')[-1].split('_')[0]


class Adesk:  # После получения доступа к данным, попробовать вытащить контрагента
	def __init__(self, api):
		# '415b6479c8df4d619ff3e957e6a262242f5c3fa6024744c2ac1ff532e8d76a1e'
		self.api = api
		self.data = {}
		self.request = f'https://api.adesk.ru/v1/transactions?api_token={self.api}'

		response = requests.get(self.request)
		while not response.ok:
			response = requests.get(self.request)
			time.sleep(1)
		self.income_json = response.json()

	def get_projects(self):
		response = requests.get(self.request)
		while not response.ok:
			response = requests.get(self.request)
			time.sleep(1)
		data_json = response.json()
		for i in range(len(data_json['projects'])):
			pprint.pprint(data_json)
			numbers_of_project = data_json['projects'][i]['name'][:4]
			id_of_project = data_json['projects'][i]['id']
			incomes = float(data_json['projects'][i]['income'])
			plan_incomes = float(data_json['projects'][i]['planIncome'])
			self.data[numbers_of_project] = (incomes, id_of_project, plan_incomes)
		return self.data

	def get_income_of_project_in_month(self, month, year, id_proj):
		income = 0
		start = datetime.datetime.strptime(f"1.{months.index(month) + 1}.{year}", "%d.%m.%Y")
		end = datetime.datetime.strptime(f"31.{months.index(month) + 1}.{year}", "%d.%m.%Y")
		for i in range(self.income_json['recordsTotal']):
			now = datetime.datetime.strptime(self.income_json['transactions'][i]['date'], "%d.%m.%Y")
			if (self.income_json['transactions'][i]['project']['id'] == id_proj) and (start <= now <= end):
				income += float(self.income_json['transactions'][i]['amount'])
		return income


class Accountant:
	def calculating_wages(self, worked_hours, standard_hours, salary_rate):
		return float(f"{(worked_hours / standard_hours) * salary_rate:.{2}f}")
